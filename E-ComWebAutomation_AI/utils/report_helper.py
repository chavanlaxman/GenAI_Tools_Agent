"""Helper functions for test reporting and artifacts."""
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

def create_excel_report(results, run_dir, timestamp):
    """
    Create detailed Excel report from test results.
    Args:
        results (list): List of test result dictionaries
        run_dir (str): Directory for test artifacts
        timestamp (str): Test run timestamp
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Test Results'

    # Define styles
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    pass_fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    fail_fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    skip_fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
    
    # Headers
    headers = ['Test Name', 'Test ID', 'Status', 'Duration (s)', 'Error Message', 'Screenshot']
    ws.append(headers)
    
    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
    
    # Add results
    for r in results:
        status = r.get('outcome', '')
        row = [
            r.get('name', ''),
            r.get('nodeid', ''),
            status,
            r.get('duration', ''),
            r.get('error_message', ''),
            r.get('screenshot', '')
        ]
        ws.append(row)
        
        # Style result row
        row_num = ws.max_row
        fill = pass_fill if status == 'passed' else fail_fill if status == 'failed' else skip_fill
        for cell in ws[row_num]:
            cell.fill = fill
    
    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    # Save report
    report_path = os.path.join(run_dir, f"test_results_{timestamp}.xlsx")
    wb.save(report_path)
    return report_path

def cleanup_old_reports(reports_dir, max_age_days=7):
    """
    Clean up old test reports to manage disk space.
    Args:
        reports_dir (str): Directory containing reports
        max_age_days (int): Maximum age of reports to keep
    """
    cutoff = datetime.now().timestamp() - (max_age_days * 86400)
    
    for item in os.listdir(reports_dir):
        item_path = os.path.join(reports_dir, item)
        if os.path.isfile(item_path):
            if os.path.getctime(item_path) < cutoff:
                try:
                    os.remove(item_path)
                except Exception as e:
                    print(f"Failed to remove old report {item}: {e}")

def consolidate_run_logs(run_dir):
    """
    Consolidate all log files from a test run into a single file.
    Args:
        run_dir (str): Test run directory containing logs
    """
    logs_dir = os.path.join(run_dir, 'logs')
    consolidated_log = os.path.join(run_dir, 'consolidated_test_run.log')
    
    with open(consolidated_log, 'w') as outfile:
        for filename in sorted(os.listdir(logs_dir)):
            if filename.endswith('.log'):
                log_path = os.path.join(logs_dir, filename)
                outfile.write(f"\n=== {filename} ===\n")
                with open(log_path) as infile:
                    outfile.write(infile.read())